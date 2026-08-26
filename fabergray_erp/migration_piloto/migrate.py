"""Producción: migración de Clientes y Productos legados (Access, vía Excel)
hacia Customer/Item. Ver README.md de este módulo para el modo de uso.
"""

import json

import frappe
from frappe.database import savepoint

from fabergray_erp.migration_piloto.importer import (
    compute_duplicate_codes,
    has_valid_price,
    import_customer,
    import_item,
    norm,
    upsert_item_price,
    validate_prerequisites,
)

DEFAULT_COMMIT_EVERY = 250
DEFAULT_PROGRESS_EVERY = 250


def _load_xlsx(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [str(h).strip() if h is not None else f"__col{i}" for i, h in enumerate(header)]
    data = []
    for r in rows:
        if all(c is None for c in r):
            continue
        data.append(dict(zip(header, r)))
    return data


def _is_omittable_customer(row):
    """Customers we refuse to create because we would have to invent
    customer_name (mandatory, never fabricated). Data-driven, not a
    hardcoded id list: any row whose Nombre is blank."""
    return not norm(row.get("Nombre"))


def _dry_run(clientes, productos, duplicate_codes):
    existing_cliente_ids = set(
        frappe.get_all("Customer", filters={"access_id_cliente": ["is", "set"]}, pluck="access_id_cliente")
    )
    existing_item_ids = set(
        frappe.get_all("Item", filters={"access_id_producto": ["is", "set"]}, pluck="access_id_producto")
    )

    omitted = [r for r in clientes if _is_omittable_customer(r)]
    creatable_clientes = [r for r in clientes if not _is_omittable_customer(r)]

    cli_new = sum(1 for r in creatable_clientes if str(r["IdCliente"]) not in existing_cliente_ids)
    cli_upd = sum(1 for r in creatable_clientes if str(r["IdCliente"]) in existing_cliente_ids)

    item_new = sum(1 for r in productos if str(r["IdProducto"]) not in existing_item_ids)
    item_upd = sum(1 for r in productos if str(r["IdProducto"]) in existing_item_ids)

    from fabergray_erp.migration_piloto.importer import PRICE_LIST, STOCK_UOM_FALLBACK, resolve_item_code

    price_create = 0
    price_update = 0
    sin_precio = 0
    for r in productos:
        if not has_valid_price(r):
            sin_precio += 1
            continue
        item_code = resolve_item_code(r.get("CodProducto"), str(r["IdProducto"]), duplicate_codes)
        exists = frappe.db.get_value(
            "Item Price", {"item_code": item_code, "price_list": PRICE_LIST, "uom": STOCK_UOM_FALLBACK}, "name"
        )
        if exists:
            price_update += 1
        else:
            price_create += 1

    return {
        "clientes_total": len(clientes),
        "clientes_nuevos": cli_new,
        "clientes_actualizar": cli_upd,
        "clientes_omitidos": len(omitted),
        "clientes_omitidos_detalle": [
            {"IdCliente": r.get("IdCliente"), "Documento": r.get("Documento"), "motivo": "REQUIERE CORRECCION MANUAL (Nombre vacio)"}
            for r in omitted
        ],
        "items_total": len(productos),
        "items_nuevos": item_new,
        "items_actualizar": item_upd,
        "item_prices_a_crear": price_create,
        "item_prices_a_actualizar": price_update,
        "productos_sin_precio": sin_precio,
        "clientes_disabled": sum(1 for r in creatable_clientes if not r.get("Habilitado")),
        "items_disabled": sum(1 for r in productos if not r.get("Habilitado")),
    }


def _migrate_customers(clientes, commit_every, progress_every, log):
    counters = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}
    errors = []
    total = len(clientes)
    for i, row in enumerate(clientes, start=1):
        if _is_omittable_customer(row):
            counters["skipped"] += 1
        else:
            with savepoint(catch=Exception):
                try:
                    _, created = import_customer(row)
                    counters["created" if created else "updated"] += 1
                except Exception as e:
                    counters["errors"] += 1
                    errors.append({"IdCliente": row.get("IdCliente"), "error": f"{type(e).__name__}: {e}"})
                    raise  # re-raise so the savepoint rolls back only this row

        if i % commit_every == 0 or i == total:
            frappe.db.commit()
        if i % progress_every == 0 or i == total:
            log(f"Clientes {i}/{total} -- created={counters['created']} updated={counters['updated']} skipped={counters['skipped']} errors={counters['errors']}")

    return counters, errors


def _migrate_items(productos, duplicate_codes, commit_every, progress_every, log):
    item_counters = {"created": 0, "updated": 0, "errors": 0}
    price_counters = {"created": 0, "updated": 0, "skipped_no_price": 0, "errors": 0}
    errors = []
    total = len(productos)

    for i, row in enumerate(productos, start=1):
        item_name = None
        with savepoint(catch=Exception):
            try:
                item_name, created = import_item(row, duplicate_codes)
                item_counters["created" if created else "updated"] += 1
            except Exception as e:
                item_counters["errors"] += 1
                errors.append({"IdProducto": row.get("IdProducto"), "stage": "item", "error": f"{type(e).__name__}: {e}"})
                raise

        if item_name and has_valid_price(row):
            with savepoint(catch=Exception):
                try:
                    _, price_created = upsert_item_price(item_name, row["Venta Publico"])
                    price_counters["created" if price_created else "updated"] += 1
                except Exception as e:
                    price_counters["errors"] += 1
                    errors.append({"IdProducto": row.get("IdProducto"), "stage": "item_price", "error": f"{type(e).__name__}: {e}"})
                    raise
        elif not has_valid_price(row):
            price_counters["skipped_no_price"] += 1

        if i % commit_every == 0 or i == total:
            frappe.db.commit()
        if i % progress_every == 0 or i == total:
            log(
                f"Productos {i}/{total} -- items(created={item_counters['created']} updated={item_counters['updated']} errors={item_counters['errors']}) "
                f"prices(created={price_counters['created']} updated={price_counters['updated']} sin_precio={price_counters['skipped_no_price']} errors={price_counters['errors']})"
            )

    return item_counters, price_counters, errors


def migrate_access_customers_and_items(
    customers_xlsx,
    products_xlsx,
    dry_run=False,
    commit_every=DEFAULT_COMMIT_EVERY,
    progress_every=DEFAULT_PROGRESS_EVERY,
):
    """Entry point for both the analysis pass and the real import.

    dry_run=True: read-only, no frappe writes, returns projected counts.
    dry_run=False: writes via frappe.get_doc()/insert()/save() only, commits
    every `commit_every` rows (never per-row), isolates each row's failure
    with a DB savepoint so one bad record cannot roll back prior progress.

    Idempotent: Customer is located by access_id_cliente, Item by
    access_id_producto, Item Price by (item_code, price_list, uom). Running
    this twice with the same inputs must produce created=0 the second time.
    """
    validate_prerequisites()  # raises and writes nothing if a prerequisite is missing

    clientes = _load_xlsx(customers_xlsx)
    productos = _load_xlsx(products_xlsx)
    duplicate_codes = compute_duplicate_codes(productos)

    if dry_run:
        result = _dry_run(clientes, productos, duplicate_codes)
        print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
        return result

    def log(msg):
        print(msg, flush=True)

    log(f"=== INICIO migracion === clientes={len(clientes)} productos={len(productos)}")
    cli_counters, cli_errors = _migrate_customers(clientes, commit_every, progress_every, log)
    log(f"=== Clientes terminado === {cli_counters}")
    item_counters, price_counters, item_errors = _migrate_items(productos, duplicate_codes, commit_every, progress_every, log)
    log(f"=== Productos/Precios terminado === items={item_counters} prices={price_counters}")

    result = {
        "customers": cli_counters,
        "items": item_counters,
        "item_prices": price_counters,
        "customer_errors": cli_errors,
        "item_errors": item_errors,
    }
    print("\n=== RESULTADO FINAL ===")
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    return result
