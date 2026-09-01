# -*- coding: utf-8 -*-
"""Fase 2 de la migración Access -> Customer: Address/Contact/Contact
Phone/Contact Email para los Customer que Fase 1 (importer.py/migrate.py)
ya creó. Ver README.md de este módulo.

No crea ni reimporta Customer -- todo Customer objetivo ya debe existir
con `access_id_cliente` (Fase 1). No toca Item/Item Price/inventario ni
ningún dato de ventas/facturación/cuentas. No geocodifica -- `country`
guarda un nombre de país, nunca coordenadas, y solo se ejecuta si el
llamador aprueba explícitamente un `country_fallback` (nunca inventado
aquí por defecto).

Arquitectura de datos (exclusivamente nativa Frappe/ERPNext, cero campos
nuevos en Customer/Address/Contact):

    Customer
     |- Address  (Dynamic Link: link_doctype="Customer", link_name=Customer.name)
     `- Contact  (Dynamic Link: link_doctype="Customer", link_name=Customer.name)
          |- Contact Phone (child table)
          `- Contact Email (child table)

Identidad de coincidencia (ver match_customer(), estrictamente en este
orden -- nunca fuzzy):
    A. Excel.IdCliente  -> Customer.access_id_cliente (match principal)
    B. Excel.Documento  -> Customer.tax_id            (fallback, compare-only)
    C. Documento ambiguo -> desambiguado por Excel.Nombre vs
       Customer.customer_name (compare-only), solo si queda EXACTAMENTE
       un candidato
    D. unmatched / E. ambiguous -- nunca escritos, solo reportados

Idempotencia (sin frappe.session.user, sin ignore_permissions, sin
insert(ignore_if_duplicate=True) como sustituto de una identidad
explícita -- section 7 del brief de Fase 2):
    - Address: se busca, por Customer, un Address ya vinculado (Dynamic
      Link) cuyo address_line1 normalizado coincida (trim/uppercase/
      collapse-whitespace/sin-acentos, compare-only). Direcciones
      distintas para el mismo Customer SÍ coexisten; nunca se deduplica
      entre Customers distintos.
    - Contact: si el Customer ya tiene CUALQUIER Contact vinculado
      (Dynamic Link), se considera ya migrado y no se toca -- no se
      actualiza ni se crea uno nuevo. Esto evita mutar un Contact que
      pudo haberse creado por otra vía (manual, u otro proceso), y hace
      la identidad tan simple y explícita como "el Customer ya tiene
      contacto" (Section 5/7 del brief: un Excel row = un Customer = a
      lo sumo un Contact migrado por este módulo).
    - Contact Phone/Email: deduplicados dentro del mismo Contact por
      clave de comparación (ver phone_compare_key()/clean_valid_email()),
      nunca por posición ni por conteo.

Bloqueos conocidos y NO resueltos silenciosamente aquí (ver auditoría de
Fase 2 -- fabergray.local, Excel real: Ciudad es 100% IDs numéricos de
Access, no nombres reales; País está 100% vacío):
    - Address.city y Address.country son REQUERIDOS nativamente
      (Address.json, reqd=1, sin tocar -- una relajación vía Property
      Setter se probó en un intento anterior de Fase 2B y fue revertida
      por decisión explícita: no se quiere city vacío). Sin un nombre de
      ciudad resoluble (Excel.Ciudad ya es un nombre real, o aparece en
      `city_id_map`, o hay un `city_fallback` explícito) o sin país
      (Excel.País, o `country_fallback` explícito), ese Address NO se
      crea -- se cuenta como bloqueado (`addresses_blocked_missing_city`/
      `addresses_blocked_missing_country`), nunca se inventa un valor
      por decisión propia de este módulo. `city_fallback`/
      `country_fallback` son exactamente eso -- un valor que el
      LLAMADOR aprueba explícitamente para esta migración puntual
      (p.ej. `city_fallback="Bucaramanga"`, `country_fallback="Colombia"`
      mientras no exista una tabla real IdCiudad->nombre), nunca
      inferidos ni asumidos por defecto aquí.
"""

import re
import unicodedata
from collections import defaultdict

import frappe

DEFAULT_ADDRESS_TYPE = "Billing"

MAX_SAMPLE = 20


# ---------------------------------------------------------------------------
# Normalización -- exclusivamente para comparar, nunca para lo que se
# guarda. Ninguna de estas funciones transforma destructivamente un valor
# que vaya a persistirse.
# ---------------------------------------------------------------------------


def norm(value):
    return "" if value is None else str(value).strip()


def normalize_id_value(value):
    """Excel/openpyxl a veces entrega una columna de ID entero como float
    (123 -> 123.0) según el formato de la celda; compare-only, colapsa eso
    de vuelta al entero en texto. Nunca toca lo que ya está guardado en
    Customer.access_id_cliente (Fase 1 ya lo escribió como texto plano;
    esto solo afecta cómo se compara un valor de Excel contra eso)."""
    s = norm(value)
    if not s:
        return ""
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return s


def normalize_tax_id(value):
    """trim, uppercase, remover todo carácter no alfanumérico -- compare
    only, según el fallback de la sección 3.B del brief."""
    s = norm(value).upper()
    return re.sub(r"[^A-Z0-9]", "", s)


def normalize_for_compare(value):
    """trim, uppercase, collapse whitespace, sin acentos (Unicode) --
    compare-only, compartido por la desambiguación por nombre (3.C) y la
    idempotencia de Address (4)."""
    s = norm(value)
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    return re.sub(r"\s+", " ", s).strip()


def normalize_phone_for_storage(value):
    """La ÚNICA transformación aplicada a un teléfono antes de guardarlo:
    trim + colapsar espacios internos redundantes. Nunca elimina
    puntuación, nunca agrega indicativo de país, nunca elimina una
    extensión legítima."""
    s = norm(value)
    if not s:
        return ""
    return re.sub(r"\s+", " ", s)


def phone_compare_key(value):
    """Clave compare-only para detectar duplicados dentro del mismo
    Contact (p.ej. Telefono1 == Celular) -- quita todo lo que no sea
    letra/dígito, así "(315)-2273268" y "315 2273268" se reconocen como
    el mismo número sin alterar jamás el valor guardado."""
    return re.sub(r"[^A-Za-z0-9]", "", norm(value)).upper()


def is_valid_phone_value(value):
    """El MISMO validador que Frappe aplica nativamente a Contact Phone.
    phone (Data, options="Phone") al guardar --
    frappe.utils.validate_phone_number()'s own PHONE_NUMBER_PATTERN regex
    (r"([0-9\ \+\_\-\,\.\*\#\(\)]){1,20}$", confirmado en
    frappe/model/base_document.py::_validate_data_fields(), rama
    data_field_options == "Phone") -- reusado tal cual, nunca una
    expresión regular propia más permisiva que Frappe luego rechazaría.
    "EXT 138" falla porque contiene letras, que el patrón nativo no
    admite."""
    if not value:
        return False
    return bool(frappe.utils.validate_phone_number(value, throw=False))


def is_numeric_city_value(value):
    """True para un ID crudo de Access ('1', '49', '9.0'), no un nombre
    real de ciudad -- exactamente la señal que pide la auditoría."""
    s = norm(value)
    if not s:
        return False
    return bool(re.fullmatch(r"\d+(\.0+)?", s))


def clean_valid_email(value):
    """Devuelve el email (trimmed) SOLO si el validador nativo de frappe
    (frappe.utils.validate_email_address) lo acepta SIN alterarlo --
    deliberadamente más estricto que "el validador logró extraer algo": si
    el parser nativo tuvo que modificar/recortar algo para considerarlo
    válido (varias direcciones en una celda, un nombre de display suelto),
    esta migración lo trata como "obviamente inválido" en vez de aceptar
    en silencio un valor corregido -- "no corregir automáticamente"
    (sección 5), aplicado literalmente. None si no es válido."""
    s = norm(value)
    if not s:
        return None
    cleaned = frappe.utils.validate_email_address(s, throw=False)
    if cleaned and cleaned == s:
        return cleaned
    return None


# ---------------------------------------------------------------------------
# Índice de Customers (una sola lectura, en memoria) + matching
# ---------------------------------------------------------------------------


def load_customer_index():
    """Read-only. Una sola llamada frappe.get_all() -- ningún match hace
    una query por fila."""
    customers = frappe.get_all(
        "Customer", fields=["name", "access_id_cliente", "tax_id", "customer_name"]
    )
    by_access_id = {}
    by_tax_id = defaultdict(list)
    for c in customers:
        aid = normalize_id_value(c.access_id_cliente)
        if aid and aid not in by_access_id:
            by_access_id[aid] = c
        tid = normalize_tax_id(c.tax_id)
        if tid:
            by_tax_id[tid].append(c)
    return by_access_id, by_tax_id


def match_customer(row, by_access_id, by_tax_id):
    """Devuelve (customer_row_o_None, match_type) con match_type en
    {"access_id", "tax_id", "tax_id_and_name", "unmatched", "ambiguous"}.
    Orden estricto de la sección 3 -- nunca continúa después de un match
    exitoso, nunca adivina."""
    id_cliente = normalize_id_value(row.get("IdCliente"))
    if id_cliente:
        customer = by_access_id.get(id_cliente)
        if customer:
            return customer, "access_id"

    documento = normalize_tax_id(row.get("Documento"))
    if documento:
        candidates = by_tax_id.get(documento, [])
        if len(candidates) == 1:
            return candidates[0], "tax_id"
        if len(candidates) > 1:
            excel_name = normalize_for_compare(row.get("Nombre"))
            name_matches = [c for c in candidates if normalize_for_compare(c.customer_name) == excel_name]
            if len(name_matches) == 1:
                return name_matches[0], "tax_id_and_name"
            return None, "ambiguous"

    return None, "unmatched"


# ---------------------------------------------------------------------------
# Address
# ---------------------------------------------------------------------------


def row_has_address(row):
    return bool(norm(row.get("Dirección")))


def existing_addresses_for_customer(customer_name):
    links = frappe.get_all(
        "Dynamic Link",
        filters={"link_doctype": "Customer", "link_name": customer_name, "parenttype": "Address"},
        pluck="parent",
    )
    if not links:
        return []
    return frappe.get_all("Address", filters={"name": ["in", links]}, fields=["name", "address_line1"])


def find_matching_address(customer_name, address_line1, cache):
    """cache: dict mutable {customer_name: [Address rows]} para no
    reconsultar el mismo Customer más de una vez en una misma pasada."""
    key_norm = normalize_for_compare(address_line1)
    if customer_name not in cache:
        cache[customer_name] = existing_addresses_for_customer(customer_name)
    for addr in cache[customer_name]:
        if normalize_for_compare(addr.address_line1) == key_norm:
            return addr.name
    return None


def resolve_city(row, city_id_map, city_fallback=None):
    """Orden estricto (aprobado explícitamente, nunca fuzzy):
    A. Un valor ya no-numérico en Excel.Ciudad se usa tal cual (es un
       nombre real) -- `city_fallback` ni se considera en este caso.
    B. Un valor numérico (ID de Access) se resuelve si aparece en
       `city_id_map` (tabla de equivalencias aprobada, opcional).
    C. Si Excel.Ciudad está vacía, o es un ID numérico SIN entrada en
       `city_id_map`: se usa `city_fallback` -- solo si el llamador lo
       pasó explícitamente (nunca asumido por este módulo, nunca
       inferido desde barrio/teléfono/dirección/nombre del cliente).
    D. Si tampoco hay `city_fallback`: None -- comportamiento seguro,
       nunca se inventa un valor y el Address correspondiente no se
       crea (ver _create_address()/addresses_blocked_missing_city)."""
    raw = norm(row.get("Ciudad"))
    if raw and not is_numeric_city_value(raw):
        return raw
    if raw:
        mapped = (city_id_map or {}).get(normalize_id_value(raw))
        if mapped:
            return mapped
    return city_fallback


def resolve_country(row, country_fallback):
    """None si no hay país resoluble. Excel.País nunca está vacío-pero-
    usado-de-otra-forma; `country_fallback` solo se aplica si el llamador
    lo pasó explícitamente (nunca asumido por este módulo)."""
    raw = norm(row.get("País"))
    if raw:
        return raw
    return country_fallback


def address_title_for(row, customer_row):
    nombre_comercial = norm(row.get("Nombre Comercial"))
    if nombre_comercial:
        return nombre_comercial
    return customer_row.customer_name


# ---------------------------------------------------------------------------
# Contact
# ---------------------------------------------------------------------------


def row_has_contact_data(row):
    return bool(
        norm(row.get("Nombre contacto"))
        or norm(row.get("Telefono1"))
        or norm(row.get("Telefono2"))
        or norm(row.get("Celular"))
        or norm(row.get("Cargo contacto"))
        or clean_valid_email(row.get("Dirección correo"))
        or clean_valid_email(row.get("DirCorreoElectrónico2"))
    )


def existing_contact_name_for_customer(customer_name):
    """The Contact already linked to this Customer via Dynamic Link, if
    any (name of the first one found -- by design this migration never
    creates more than one per Customer, see the module's own idempotency
    note above). None if the Customer has no Contact yet."""
    links = frappe.get_all(
        "Dynamic Link",
        filters={"link_doctype": "Customer", "link_name": customer_name, "parenttype": "Contact"},
        pluck="parent",
        limit_page_length=1,
    )
    return links[0] if links else None


def customer_already_has_contact(customer_name):
    return existing_contact_name_for_customer(customer_name) is not None


def collect_phones(row):
    """Devuelve [(cleaned_value, is_mobile)] -- SOLO teléfonos que pasan
    is_valid_phone_value() (la misma validación nativa que Frappe aplica
    al guardar), deduplicado por phone_compare_key(), preservando el
    orden Telefono1 -> Telefono2 -> Celular. Si Telefono1 == Celular (u
    otra combinación), solo la primera aparición sobrevive. Un valor que
    no pasa la validación (p.ej. "EXT 138") nunca aparece aquí -- ver
    invalid_phones_in_row() para esos; nunca se descarta toda la fila por
    esto, cada teléfono se valida de forma independiente."""
    candidates = [
        (row.get("Telefono1"), False),
        (row.get("Telefono2"), False),
        (row.get("Celular"), True),
    ]
    seen_keys = set()
    result = []
    for raw, is_mobile in candidates:
        cleaned = normalize_phone_for_storage(raw)
        if not cleaned or not is_valid_phone_value(cleaned):
            continue
        key = phone_compare_key(cleaned)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        result.append((cleaned, is_mobile))
    return result


def invalid_phones_in_row(row):
    """Valores de Telefono1/Telefono2/Celular que NO pasan
    is_valid_phone_value() -- p.ej. "EXT 138" (solo una extensión, sin
    número base). Nunca insertados como Contact Phone, nunca
    concatenados con otro número, nunca guardados en notas -- el
    llamador es responsable de contarlos (invalid_phones_skipped)."""
    invalid = []
    for raw in (row.get("Telefono1"), row.get("Telefono2"), row.get("Celular")):
        cleaned = normalize_phone_for_storage(raw)
        if cleaned and not is_valid_phone_value(cleaned):
            invalid.append(cleaned)
    return invalid


def row_has_usable_contact_data(row):
    """Distinto de row_has_contact_data() arriba (esa es la señal
    cruda/informativa que usa el reporte de dry-run): esta es la señal
    REAL que _create_contact() usa para decidir si vale la pena crear el
    documento, evaluada DESPUÉS de filtrar teléfonos/emails que Frappe
    rechazaría. Nombre/Cargo contacto no necesitan filtro (texto libre);
    un Telefono1="EXT 138" sin nada más NO cuenta como dato usable."""
    return bool(
        norm(row.get("Nombre contacto"))
        or norm(row.get("Cargo contacto"))
        or collect_phones(row)
        or collect_emails(row)
    )


def collect_emails(row):
    """Devuelve [email, ...] válidos y deduplicados (compare-only por
    valor exacto ya limpio), preservando el orden Dirección correo ->
    DirCorreoElectrónico2. Emails inválidos se descartan aquí -- el
    llamador es responsable de registrarlos en el reporte."""
    result = []
    seen = set()
    for raw in (row.get("Dirección correo"), row.get("DirCorreoElectrónico2")):
        cleaned = clean_valid_email(raw)
        if not cleaned:
            continue
        key = cleaned.upper()
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def invalid_emails_in_row(row):
    invalid = []
    for raw in (row.get("Dirección correo"), row.get("DirCorreoElectrónico2")):
        raw_norm = norm(raw)
        if raw_norm and not clean_valid_email(raw_norm):
            invalid.append(raw_norm)
    return invalid


# ---------------------------------------------------------------------------
# Carga de Excel (idéntico contrato a migrate.py::_load_xlsx)
# ---------------------------------------------------------------------------


def load_xlsx(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [str(h).strip() if h is not None else f"__col{i}" for i, h in enumerate(header)]
    data = []
    for r in rows:
        if all(c is None for c in r):
            continue
        data.append(dict(zip(header, r)))
    return data


# ---------------------------------------------------------------------------
# Auditoría / Dry run -- 100% read-only. Sin frappe.db.commit(), sin
# save()/insert()/db_set()/set_value()/delete(), sin SQL de escritura.
# ---------------------------------------------------------------------------


def audit_customer_contacts(excel_path):
    """Alias legible de migrate_customer_contacts(excel_path, dry_run=True)
    -- mismo resultado, nombre explícito para uso puramente de auditoría."""
    return migrate_customer_contacts(excel_path, dry_run=True)


def _dry_run_report(
    rows, by_access_id, by_tax_id, country_fallback, city_id_map,
    migrate_addresses=True, migrate_contacts=True, city_fallback=None,
):
    report = {
        "excel_rows": len(rows),
        "customers_matched_by_access_id": 0,
        "customers_matched_by_tax_id": 0,
        "customers_matched_by_tax_id_and_name": 0,
        "unmatched_rows": 0,
        "ambiguous_rows": 0,
        "rows_with_address": 0,
        "addresses_would_create": 0,
        "addresses_already_exist": 0,
        "addresses_duplicate_in_excel": 0,
        "rows_without_address": 0,
        "addresses_blocked_missing_city": 0,
        "addresses_blocked_missing_country": 0,
        "rows_with_contact_data": 0,
        "contacts_would_create": 0,
        "contacts_already_exist": 0,
        "phone1_count": 0,
        "phone2_count": 0,
        "mobile_count": 0,
        "unique_phones_would_create": 0,
        "email1_count": 0,
        "email2_count": 0,
        "valid_emails": 0,
        "invalid_emails": 0,
        "city_values_count": 0,
        "numeric_city_values_count": 0,
        "country_empty_count": 0,
    }
    samples = {"unmatched": [], "ambiguous": [], "invalid_emails": [], "numeric_cities": []}

    address_cache = {}
    # (customer_name, address_line1_norm) ya "reclamado" en ESTA pasada,
    # sea porque ya existía en BD o porque una fila anterior de este mismo
    # Excel ya lo contó como "a crear" -- distingue "ya existía" de
    # "duplicado dentro del propio Excel".
    claimed_addresses = set()
    contact_already_seen_customers = set()

    for row in rows:
        customer, match_type = match_customer(row, by_access_id, by_tax_id)

        if match_type == "access_id":
            report["customers_matched_by_access_id"] += 1
        elif match_type == "tax_id":
            report["customers_matched_by_tax_id"] += 1
        elif match_type == "tax_id_and_name":
            report["customers_matched_by_tax_id_and_name"] += 1
        elif match_type == "ambiguous":
            report["ambiguous_rows"] += 1
            if len(samples["ambiguous"]) < MAX_SAMPLE:
                samples["ambiguous"].append(
                    {"IdCliente": row.get("IdCliente"), "Documento": row.get("Documento"), "Nombre": row.get("Nombre")}
                )
        else:
            report["unmatched_rows"] += 1
            if len(samples["unmatched"]) < MAX_SAMPLE:
                samples["unmatched"].append(
                    {"IdCliente": row.get("IdCliente"), "Documento": row.get("Documento"), "Nombre": row.get("Nombre")}
                )

        ciudad_raw = norm(row.get("Ciudad"))
        if ciudad_raw:
            report["city_values_count"] += 1
            if is_numeric_city_value(ciudad_raw):
                report["numeric_city_values_count"] += 1
                if len(samples["numeric_cities"]) < MAX_SAMPLE:
                    samples["numeric_cities"].append(ciudad_raw)

        # -- Address --------------------------------------------------------
        if not row_has_address(row):
            report["rows_without_address"] += 1
        else:
            report["rows_with_address"] += 1
            if not norm(row.get("País")):
                report["country_empty_count"] += 1

            if migrate_addresses and customer:
                address_line1 = norm(row.get("Dirección"))
                claim_key = (customer.name, normalize_for_compare(address_line1))
                existing = find_matching_address(customer.name, address_line1, address_cache)
                if existing:
                    report["addresses_already_exist"] += 1
                elif claim_key in claimed_addresses:
                    report["addresses_duplicate_in_excel"] += 1
                else:
                    claimed_addresses.add(claim_key)
                    country = resolve_country(row, country_fallback)
                    city = resolve_city(row, city_id_map, city_fallback)
                    if not country:
                        report["addresses_blocked_missing_country"] += 1
                    elif not city:
                        report["addresses_blocked_missing_city"] += 1
                    else:
                        report["addresses_would_create"] += 1

        # -- Contact ----------------------------------------------------------
        has_contact_data = row_has_contact_data(row)
        if has_contact_data:
            report["rows_with_contact_data"] += 1
            if migrate_contacts and customer:
                if customer.name in contact_already_seen_customers or customer_already_has_contact(customer.name):
                    report["contacts_already_exist"] += 1
                else:
                    contact_already_seen_customers.add(customer.name)
                    report["contacts_would_create"] += 1

                    phones = collect_phones(row)
                    report["unique_phones_would_create"] += len(phones)

        if norm(row.get("Telefono1")):
            report["phone1_count"] += 1
        if norm(row.get("Telefono2")):
            report["phone2_count"] += 1
        if norm(row.get("Celular")):
            report["mobile_count"] += 1

        if norm(row.get("Dirección correo")):
            report["email1_count"] += 1
        if norm(row.get("DirCorreoElectrónico2")):
            report["email2_count"] += 1
        report["valid_emails"] += len(collect_emails(row))
        for invalid in invalid_emails_in_row(row):
            report["invalid_emails"] += 1
            if len(samples["invalid_emails"]) < MAX_SAMPLE:
                samples["invalid_emails"].append(invalid)

    report["samples"] = samples
    return report


# ---------------------------------------------------------------------------
# Migración real -- solo se ejecuta escritura cuando dry_run=False.
# ---------------------------------------------------------------------------


def _create_address(customer, row, country_fallback, city_id_map, city_fallback=None):
    """Devuelve (doc, outcome) con outcome en {"created",
    "blocked_missing_city", "blocked_missing_country"}.

    Address.city SIGUE siendo obligatorio nativamente (Meta sin tocar --
    la relajación vía Property Setter que se probó en un intento anterior
    de Fase 2B fue revertida por decisión explícita: no queremos city
    vacío). Sin un nombre de ciudad resoluble (ver resolve_city() -- ni
    mapping en city_id_map ni city_fallback explícito), el Address NO se
    crea -- se cuenta como bloqueado, nunca se inventa un valor por
    defecto propio de esta función.

    country también sigue siendo obligatorio -- sin uno resoluble
    (Excel.País vacío y sin country_fallback explícito), el Address
    tampoco se crea."""
    country = resolve_country(row, country_fallback)
    if not country:
        return None, "blocked_missing_country"

    city = resolve_city(row, city_id_map, city_fallback)
    if not city:
        return None, "blocked_missing_city"

    doc = frappe.new_doc("Address")
    doc.address_title = address_title_for(row, customer)
    doc.address_type = DEFAULT_ADDRESS_TYPE
    doc.address_line1 = norm(row.get("Dirección"))
    doc.city = city
    doc.country = country
    doc.append("links", {"link_doctype": "Customer", "link_name": customer.name})
    doc.insert()
    return doc, "created"


def _create_contact(customer, row):
    """Devuelve el Contact insertado, o None si -- después de filtrar
    teléfonos/emails inválidos -- la fila no tiene ningún dato usable
    real (nunca inserta un Contact vacío). El llamador cuenta un None
    como contact_rows_skipped_no_valid_data, nunca como error."""
    if not row_has_usable_contact_data(row):
        return None

    doc = frappe.new_doc("Contact")
    nombre_contacto = norm(row.get("Nombre contacto"))
    if nombre_contacto:
        doc.first_name = nombre_contacto  # texto completo, sin inventar apellido (sección 5)
    doc.company_name = customer.customer_name
    designation = norm(row.get("Cargo contacto"))
    if designation:
        doc.designation = designation
    for phone, is_mobile in collect_phones(row):
        doc.append(
            "phone_nos",
            {"phone": phone, "is_primary_mobile_no": 1 if is_mobile else 0},
        )
    for i, email in enumerate(collect_emails(row)):
        doc.append("email_ids", {"email_id": email, "is_primary": 1 if i == 0 else 0})
    doc.append("links", {"link_doctype": "Customer", "link_name": customer.name})
    doc.insert()
    return doc


def _maybe_set_primary_address(customer_name, address_name):
    """Solo si el Customer no tiene ya un primary address -- nunca
    sobrescribe uno existente (sección 6)."""
    current = frappe.db.get_value("Customer", customer_name, "customer_primary_address")
    if current:
        return
    frappe.db.set_value("Address", address_name, "is_primary_address", 1)
    frappe.db.set_value("Customer", customer_name, "customer_primary_address", address_name)


def _maybe_set_primary_contact(customer_name, contact_name):
    """Returns True if it just SET customer_primary_contact (was empty),
    False if the Customer already had one and it was left untouched
    (sección 6 -- "no sobrescribirlo")."""
    current = frappe.db.get_value("Customer", customer_name, "customer_primary_contact")
    if current:
        return False
    frappe.db.set_value("Contact", contact_name, "is_primary_contact", 1)
    frappe.db.set_value("Customer", customer_name, "customer_primary_contact", contact_name)
    return True


def _migrate_real(
    rows,
    by_access_id,
    by_tax_id,
    country_fallback,
    city_id_map,
    commit_every,
    progress_every,
    log,
    migrate_addresses=True,
    migrate_contacts=True,
    city_fallback=None,
):
    from frappe.database import savepoint

    counters = {
        "unmatched": 0,
        "ambiguous": 0,
        "addresses_created": 0,
        "addresses_already_exist": 0,
        "addresses_blocked_missing_city": 0,
        "addresses_blocked_missing_country": 0,
        "contacts_created": 0,
        "contacts_already_exist": 0,
        "phones_created": 0,
        "emails_created": 0,
        "emails_invalid_skipped": 0,
        "invalid_phones_skipped": 0,
        "contacts_recovered_from_invalid_phone_rows": 0,
        "contact_rows_skipped_no_valid_data": 0,
        "dynamic_links_created": 0,
        "primary_contacts_set": 0,
        "primary_contacts_preserved": 0,
        "errors": 0,
    }
    errors = []
    address_cache = {}
    total = len(rows)

    for i, row in enumerate(rows, start=1):
        customer, match_type = match_customer(row, by_access_id, by_tax_id)
        if match_type == "unmatched":
            counters["unmatched"] += 1
        elif match_type == "ambiguous":
            counters["ambiguous"] += 1

        if customer:
            # migrate_addresses=False guarantees this entire block never
            # runs: no Address created/modified, no Dynamic Link for
            # Address, customer_primary_address never touched,
            # country_fallback/city_id_map never even read.
            if migrate_addresses and row_has_address(row):
                address_line1 = norm(row.get("Dirección"))
                existing = find_matching_address(customer.name, address_line1, address_cache)
                if existing:
                    counters["addresses_already_exist"] += 1
                else:
                    with savepoint(catch=Exception):
                        try:
                            doc, outcome = _create_address(
                                customer, row, country_fallback, city_id_map, city_fallback
                            )
                            if outcome == "created":
                                counters["addresses_created"] += 1
                                address_cache.setdefault(customer.name, []).append(
                                    frappe._dict(name=doc.name, address_line1=doc.address_line1)
                                )
                                _maybe_set_primary_address(customer.name, doc.name)
                            else:
                                counters[f"addresses_{outcome}"] += 1
                        except Exception as e:
                            counters["errors"] += 1
                            errors.append(
                                {"IdCliente": row.get("IdCliente"), "stage": "address", "error": f"{type(e).__name__}: {e}"}
                            )
                            raise

            if migrate_contacts and row_has_contact_data(row):
                # Reset every row, on purpose -- a stale contact_name left
                # over from a PRIOR row's successful creation must never
                # leak into this row's own primary-contact assignment if
                # this row's own _create_contact() below fails/skips and
                # the savepoint swallows the exception (catch=Exception
                # logs and rolls back the row, it does not re-raise out of
                # the `with` block, so falling through to _maybe_set_
                # primary_contact() without this reset would silently
                # assign a DIFFERENT customer's own Contact as this one's
                # primary -- this is the exact bug found and fixed live
                # against fabergray.local, kept here as the guard, not
                # just as a comment).
                contact_name = None
                existing_contact_name = existing_contact_name_for_customer(customer.name)
                if existing_contact_name:
                    counters["contacts_already_exist"] += 1
                    contact_name = existing_contact_name
                else:
                    invalid_phones = invalid_phones_in_row(row)
                    counters["invalid_phones_skipped"] += len(invalid_phones)
                    with savepoint(catch=Exception):
                        try:
                            contact = _create_contact(customer, row)
                            if contact is None:
                                # Every phone/email this row had was
                                # invalid, and there is no name/designation
                                # either -- nothing usable survived
                                # filtering. Never insert an empty Contact.
                                counters["contact_rows_skipped_no_valid_data"] += 1
                            else:
                                counters["contacts_created"] += 1
                                counters["phones_created"] += len(contact.phone_nos)
                                counters["emails_created"] += len(contact.email_ids)
                                counters["dynamic_links_created"] += 1
                                if invalid_phones:
                                    counters["contacts_recovered_from_invalid_phone_rows"] += 1
                                contact_name = contact.name
                        except Exception as e:
                            counters["errors"] += 1
                            errors.append(
                                {"IdCliente": row.get("IdCliente"), "stage": "contact", "error": f"{type(e).__name__}: {e}"}
                            )
                            raise

                # contact_name stays None when this row's own contact was
                # skipped (no valid data) or genuinely failed above --
                # nothing valid to assign as primary in either case, and
                # this Customer's customer_primary_contact is never touched.
                if contact_name is not None:
                    counters["emails_invalid_skipped"] += len(invalid_emails_in_row(row))
                    if _maybe_set_primary_contact(customer.name, contact_name):
                        counters["primary_contacts_set"] += 1
                    else:
                        counters["primary_contacts_preserved"] += 1

        if i % commit_every == 0 or i == total:
            frappe.db.commit()
        if i % progress_every == 0 or i == total:
            log(f"Contactos {i}/{total} -- {counters}")

    return counters, errors


def migrate_customer_contacts(
    excel_path,
    dry_run=True,
    country_fallback=None,
    city_id_map=None,
    city_fallback=None,
    migrate_addresses=True,
    migrate_contacts=True,
    commit_every=250,
    progress_every=250,
):
    """Punto de entrada de Fase 2.

    dry_run=True (default): 100% read-only -- ni un frappe.db.commit(),
    save(), insert(), db_set(), set_value() ni delete(); devuelve el
    reporte completo de conteos (ver README/sección 9 del brief).

    dry_run=False: escribe Address/Contact/Contact Phone/Contact Email
    vía frappe.new_doc()/insert() únicamente, con savepoint por fila (una
    fila con error nunca revierte el progreso previo), commit cada
    `commit_every` filas.

    migrate_addresses/migrate_contacts (ambos True por defecto, para no
    romper el comportamiento previo): apagar `migrate_addresses` es una
    garantía dura, no una sugerencia -- con False, el bloque de Address
    completo (_create_address()/_maybe_set_primary_address()) nunca se
    alcanza; ningún Address se crea o modifica, ningún Dynamic Link de
    Address se crea, customer_primary_address nunca se toca, y
    `country_fallback`/`city_id_map`/`city_fallback` no se leen en
    absoluto (no son requeridos). Simétricamente, apagar
    `migrate_contacts` es la misma garantía dura para Contact/Contact
    Phone/Contact Email/Dynamic Link(Contact)/customer_primary_contact --
    ninguno se toca.

    Address.city/Address.country SIGUEN siendo obligatorios nativamente
    (Meta sin tocar -- una relajación vía Property Setter se probó y fue
    revertida por decisión explícita: no se quiere city vacío). Cuando
    migrate_addresses=True, resolve_city()/resolve_country() deciden en
    orden estricto, nunca fuzzy:
        ciudad: nombre real en Excel.Ciudad > city_id_map (ID numérico
                mapeado) > city_fallback explícito > bloqueado
                (addresses_blocked_missing_city)
        país:   Excel.País > country_fallback explícito > bloqueado
                (addresses_blocked_missing_country)
    Ninguno de los tres (`country_fallback`/`city_id_map`/
    `city_fallback`) se asume nunca por este módulo -- deben pasarse
    explícitamente.

    Nunca crea ni modifica un Customer -- todo Customer objetivo ya debe
    existir con access_id_cliente (Fase 1)."""
    rows = load_xlsx(excel_path)
    by_access_id, by_tax_id = load_customer_index()

    if dry_run:
        return _dry_run_report(
            rows, by_access_id, by_tax_id, country_fallback, city_id_map,
            migrate_addresses=migrate_addresses, migrate_contacts=migrate_contacts,
            city_fallback=city_fallback,
        )

    def log(msg):
        print(msg, flush=True)

    log(f"=== INICIO Fase 2 (Address/Contact) === filas={len(rows)} migrate_addresses={migrate_addresses} migrate_contacts={migrate_contacts}")
    counters, errors = _migrate_real(
        rows, by_access_id, by_tax_id, country_fallback, city_id_map, commit_every, progress_every, log,
        migrate_addresses=migrate_addresses, migrate_contacts=migrate_contacts, city_fallback=city_fallback,
    )
    log(f"=== Fase 2 terminada === {counters}")
    return {"counters": counters, "errors": errors}
